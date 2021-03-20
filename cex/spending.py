from openpyxl import load_workbook
import json

wb = load_workbook("./decile.xlsx")
ws = wb.active

start_row = 7
end_row = 789

def get_quantile_data(quantile):
  root = {}
  object_at_indent = {0: root}

  previous_indent = 0
  current_indent = 0
  current_label = None
  for row in range(start_row, end_row):
    label_cell = ws["A"+str(row)]
    label_value = label_cell.value
    label_indent = int(label_cell.alignment.indent + 2)/2
    if label_value in ["Share", "SE", "CV(%)"]:
      continue
    elif label_value:
      current_label = label_value
      new_object = {}
      if label_value == "Mean":
        current_indent += 1
        low_col = chr(65 + 2 + quantile *2)
        high_col = chr(65 + 3 + quantile *2)
        try:
          low_mean_cell_value = int(ws[low_col+str(row)].value)
        except:
          low_mean_cell_value = 0
        try:
          high_mean_cell_value = int(ws[high_col+str(row)].value)
        except:
          high_mean_cell_value = 0

        mean_value = (low_mean_cell_value + high_mean_cell_value) / 2
        new_object = mean_value
      else:
        current_indent = label_indent
      object_at_indent[current_indent-1][current_label] = new_object
      object_at_indent[current_indent] = new_object

  income_after_taxes = root["Sources of income and personal taxes:"]["Income after taxes"]["Mean"]
  average_annual_expenditures = root["Average annual expenditures"]["Mean"]
  savings = income_after_taxes - average_annual_expenditures
  taxes = root["Sources of income and personal taxes:"]["Personal taxes (contains some imputed values)"]
  return ( {"Personal Taxes" : taxes if taxes['Mean'] > 0 else {'Mean': '0'} ,
          "Savings": {"Mean": savings if savings > 0 else 0},
          "Average annual expenditures" : root["Average annual expenditures"],
          })

name_map = {"Fuel oil and other fuels": "Fuel oil, etc",
          "Gasoline, and motor oil": "Gasoline"}

def d3_ify(data, category_name):
  value = 0
  children = []
  for key in data:
    if key == "Mean":
      value = data[key]
    else:
      children.append(d3_ify(data[key], key))
  if category_name in name_map:
    category_name = name_map[category_name]
  if len(children):
    return {'name' : category_name, "children": children}
  else:
    return {'name' : category_name, "value": value }

output = []
for quantile in range(5):
  output.append(d3_ify(get_quantile_data(quantile), quantile))

print (json.dumps(output, indent=4))
